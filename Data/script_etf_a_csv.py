import argparse
import os
import pandas as pd
import openpyxl

# Control de argumentos y recogida de excels de la ruta dada de entrada
parser = argparse.ArgumentParser()
parser.add_argument('excel_path', type=str)
parser.add_argument('excel_name', type=str)
args = parser.parse_args()
all_etf_excel = [f for f in os.listdir(args.excel_path) if f.endswith('.xlsx')]

# Cargar el archivo de búsqueda y creamos diccionario
wb_search = openpyxl.load_workbook(args.excel_name)
sheet_search = wb_search.active
etf_tickers_dicc = {}
for row in range(3, 90):
    etf_name = sheet_search.cell(row=row, column=2).value
    ticker = sheet_search.cell(row=row, column=1).value
    if etf_name and ticker:
        etf_tickers_dicc[etf_name.strip()] = ticker.strip()
wb_search.close()

# Procesamos cada Excel como un DataFrame
for etf_excel in all_etf_excel:

    excel_path_etf = os.path.join(args.excel_path, etf_excel)
    df = pd.read_excel(excel_path_etf, sheet_name='Historical')
    wb = openpyxl.load_workbook(excel_path_etf)

    # Limpiamos datos y renombramos
    df = df.drop(columns=['Ex-Dividends'])
    df = df.rename(columns={'As Of': 'date', 'NAV per Share': 'nav_value', 'Shares Outstanding': 'shares'})
    df['date'] = pd.to_datetime(df['date'])

    # Guardar el valor anterior
    previous_value = None
    
    # Iterar sobre las filas y rellenar valores faltantes en la columna 'shares'
    for index, row in df.iterrows():
        if row['shares'] == "--" and previous_value is not None:
            df.at[index, 'shares'] = previous_value
        else:
            previous_value = row['shares']

    # Exportar el DataFrame a un archivo CSV
    etf_name = wb['Holdings']['A2'].value.strip()  
    etf_name_save = etf_tickers_dicc.get(etf_name, etf_name)
    df.to_csv(
        os.path.join(
            "Data/PreProcessed/etf/",
            etf_name_save + ".csv"
        ),
        index=False,
        sep=",",
        header=True,
    )
        
    print(f'Se ha exportado el archivo CSV como \'{etf_name_save}\'')

